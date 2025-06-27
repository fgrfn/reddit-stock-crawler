import os
import pickle
import pandas as pd
from openpyxl import load_workbook
import shutil

PICKLE_DIR = 'pickle'
TEMPLATE_PATH = 'crawler_results-vorlage.xlsx'
OUTPUT_PATH = 'crawler_results_aktuell.xlsx'

def read_pickle_files():
    data = []
    for file in os.listdir(PICKLE_DIR):
        if file.endswith('.pkl'):
            path = os.path.join(PICKLE_DIR, file)
            with open(path, 'rb') as f:
                data.append(pickle.load(f))
    return data

def create_dataframe(data_list):
    all_keys = set()
    for entry in data_list:
        all_keys.update(entry['results'].keys())
    all_keys = sorted(all_keys)
    rows = []
    for entry in data_list:
        row = {k: 0 for k in all_keys}
        row.update(entry['results'])
        row['run_id'] = entry['run_id']
        rows.append(row)
    df = pd.DataFrame(rows)
    df = df.sort_values(by='run_id').reset_index(drop=True)
    return df[['run_id'] + all_keys]

def update_excel_template(df):
    if not os.path.exists(TEMPLATE_PATH):
        print(f"❌ Excel-Vorlage nicht gefunden: {TEMPLATE_PATH}")
        return
    shutil.copy2(TEMPLATE_PATH, OUTPUT_PATH)
    wb = load_workbook(OUTPUT_PATH)
    ws = wb.active
    ws.delete_rows(2, ws.max_row - 1)
    for col_idx, col_name in enumerate(df.columns, 1):
        ws.cell(row=1, column=col_idx, value=col_name)
    for row_idx, row in enumerate(df.values, 2):
        for col_idx, val in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=val)
    wb.save(OUTPUT_PATH)
    print(f"✅ Excel-Datei aktualisiert: {OUTPUT_PATH}")

def main():
    print("📊 Lade Pickle-Dateien...")
    data = read_pickle_files()
    if not data:
        print("🚫 Keine Pickle-Dateien gefunden.")
        return
    print(f"🔢 {len(data)} Datensätze geladen")
    df = create_dataframe(data)
    update_excel_template(df)

if __name__ == '__main__':
    main()
